
import io

from core.extractors import extract_content
def extract_text_from_bytes(
    data: bytes,
    filename: str | None = None,
) -> str:

    extracted = extract_content(
        data=data,
        filename=filename or "unknown.bin",
    )

    return extracted.text or ""


def _extract_pdf(data: bytes) -> str:
    import pdfplumber

    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text_parts.append(t)
    return "\n".join(text_parts)


def _ocr_pdf(data: bytes) -> str:
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except Exception as e:
        print(f"⚠️ OCR PDF dependencies missing: {e}")
        return ""

    text_parts: list[str] = []
    images = convert_from_bytes(data)
    for img in images:
        t = pytesseract.image_to_string(img)
        if t:
            text_parts.append(t)
    return "\n".join(text_parts)


def _extract_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    text_parts: list[str] = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text_parts.append(str(cell))

    return "\n".join(text_parts)


def _ocr_image(data: bytes) -> str:
    try:
        from PIL import Image
        import pytesseract
    except Exception as e:
        print(f"⚠️ OCR image dependencies missing: {e}")
        return ""

    img = Image.open(io.BytesIO(data))
    return pytesseract.image_to_string(img)


