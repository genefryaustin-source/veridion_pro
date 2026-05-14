import io

from openpyxl import load_workbook

from core.extractors.base import (
    BaseExtractor,
    ExtractedContent,
)


class ExcelExtractor(BaseExtractor):

    supported_extensions = {
        ".xlsx",
        ".xls",
    }

    supported_content_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }

    def extract(
        self,
        data: bytes,
        filename: str,
        content_type: str,
        metadata=None,
    ) -> ExtractedContent:

        metadata = metadata or {}

        warnings = []

        combined_text = []

        extraction_method = "openpyxl"

        confidence = "HIGH"

        try:

            wb = load_workbook(
                filename=io.BytesIO(data),
                data_only=True,
            )

            print(
                "📊 XLSX SHEETS:",
                wb.sheetnames
            )

            for sheet_name in wb.sheetnames:

                try:

                    ws = wb[sheet_name]

                    combined_text.append(
                        f"\n\n===== SHEET: {sheet_name} =====\n\n"
                    )

                    for row in ws.iter_rows(
                        values_only=True
                    ):

                        values = []

                        for cell in row:

                            if cell is None:
                                continue

                            values.append(
                                str(cell)
                            )

                        if values:

                            combined_text.append(
                                " | ".join(values)
                            )

                except Exception as e:

                    print(
                        "❌ XLSX SHEET FAILED:",
                        sheet_name,
                        e,
                    )

                    warnings.append(
                        f"{sheet_name}: {e}"
                    )

        except Exception as e:

            print(
                "❌ XLSX EXTRACTION FAILED:",
                e
            )

            warnings.append(str(e))

            confidence = "LOW"

        final_text = "\n".join(
            combined_text
        )

        return ExtractedContent(

            text=final_text,

            filename=filename,

            content_type=content_type,

            extension=".xlsx",

            extraction_method=extraction_method,

            confidence=confidence,

            metadata=metadata,

            warnings=warnings,
        )